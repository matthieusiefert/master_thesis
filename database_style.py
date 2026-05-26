import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

df = pd.read_csv('SGA_table_positive_final.csv')

wb = Workbook()
ws = wb.active
ws.title = "SGA Steroidal Saponins"

# ── Colors ────────────────────────────────────────────────────────
NAVY        = "1B3A6B"   # CANOPUS_formula
NAVY_MID    = "1F4E8C"   # CANOPUS_structure
NAVY_STEEL  = "2E6DA4"   # peak areas
NAVY_SLATE  = "3B5998"   # scores/proba
NAVY_LIGHT  = "EAF1FB"   # alternating row (light blue tint)
WHITE       = "FFFFFF"
TEXT_WHITE  = "FFFFFF"
TEXT_DARK   = "1A1A1A"

def group_color(col: str) -> str:
    cl = col.lower()
    if col.startswith("CANOPUS_formula"):    return NAVY
    if col.startswith("CANOPUS_structure"):  return NAVY_MID
    if col.startswith("ident_"):             return NAVY_STEEL
    if any(k in cl for k in ("confidence", "sirius", "csi", "probability")):
        return NAVY_SLATE
    return NAVY

headers = list(df.columns)

# ── Header row ─────────────────────────────────────────────────────────────────
thin = Side(style="thin", color="FFFFFF")
border = Border(left=thin, right=thin, bottom=thin)

for ci, col in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.fill      = PatternFill("solid", start_color=group_color(col))
    cell.font      = Font(name="Arial", bold=True, color=TEXT_WHITE, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border

ws.row_dimensions[1].height = 56

# ── Data rows ──────────────────────────────────────────────────────────────────
n_rows = len(df)
for ri, row in enumerate(df.itertuples(index=False), start=2):
    bg = NAVY_LIGHT if ri % 2 == 0 else WHITE
    rfill = PatternFill("solid", start_color=bg)
    for ci, value in enumerate(row, 1):
        cell = ws.cell(row=ri, column=ci, value=value)
        cell.fill      = rfill
        cell.font      = Font(name="Arial", size=9, color=TEXT_DARK)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[ri].height = 14

# ── Column widths ──────────────────────────────────────────────────────────────
LONG_TEXT   = {"smiles", "partners", "pubchemids", "GNPSLinkout_Cluster", "GNPSLinkout_Network"}
MEDIUM_TEXT = {"Compound_Name", "name", "shared name", "SIRIUS_best_annotation",
               "molecularFormula", "InChIkey2D",
               "CANOPUS_formula_ClassyFire#most specific class",
               "CANOPUS_structure_ClassyFire#most specific class"}

for ci, col in enumerate(headers, 1):
    ltr = get_column_letter(ci)
    if col in LONG_TEXT:                              ws.column_dimensions[ltr].width = 42
    elif col in MEDIUM_TEXT or col.startswith("CANOPUS"): ws.column_dimensions[ltr].width = 26
    elif col.startswith("ident_"):                    ws.column_dimensions[ltr].width = 15
    else:                                             ws.column_dimensions[ltr].width = 17

# ── Freeze + auto-filter ───────────────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── Conditional formatting: scores → gradient rouge/jaune/vert ────────────────
last = n_rows + 1
score_cols = [i for i, c in enumerate(headers, 1)
              if any(k in c.lower() for k in ("confidence", "sirius", "csi", "probability"))]
for ci in score_cols:
    ltr = get_column_letter(ci)
    ws.conditional_formatting.add(
        f"{ltr}2:{ltr}{last}",
        ColorScaleRule(
            start_type="min",            start_color="D94040",
            mid_type="percentile", mid_value=50, mid_color="F5C518",
            end_type="max",              end_color="2E7D32"
        )
    )

# ── Conditional formatting: peak areas → data bars bleu ──────────────────────
peak_cols = [i for i, c in enumerate(headers, 1) if c.startswith("ident_")]
for ci in peak_cols:
    ltr = get_column_letter(ci)
    ws.conditional_formatting.add(
        f"{ltr}2:{ltr}{last}",
        DataBarRule(start_type="min", start_value=0,
                    end_type="max",   end_value=None,
                    color="2E6DA4")
    )

out = "SGA_table_positive_styled.xlsx"
wb.save(out)
print(f"Saved → {out}  ({n_rows} rows, {len(headers)} columns)")