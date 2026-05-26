import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Paths ---
QUANT_PATH  = "MS1_Positive_Quantification_10_04_iimn_gnps_quant.csv"
ANNOT_PATH  = "MS1_Positive_Quantification_10_04_annotations.csv"
OUTPUT_PATH = "merged_quant_annotations.xlsx"

# --- MZmine parameters ---
RT_TOL = 0.10  # RT tolerance sample-to-sample (min) — threshold for true isomers vs artifacts

# --- Columns to drop (always empty) ---
DROP_COLS = ["usi", "entry_id", "iupac_name", "cas", "internal_id", "query_spectrum_usi"]

# --- Sample-group color map (hex without #) ---
GROUP_COLORS = {
    "FRA": "FFFDE0",
    "ITH": "FFE8D0",
    "SIE": "FFD8D0",
    "ILE": "D0E8FF",
    "ZAN": "E0FFFF",
    "MON": "FFFFE0",
    "UNG": "E0FFE0",
}

# --- Column width rules (Excel units) ---
# Columns not matched fall back to DEFAULT_WIDTH
COL_WIDTHS = {
    "row ID":                    6,
    "row m/z":                   9,
    "row retention time":        7,
    "row ion mobility":          7,
    "row ion mobility unit":     7,
    "row CCS":                   6,
    "correlation group ID":      7,
    "annotation network number": 7,
    "best ion":                  8,
    "auto MS2 verify":           7,
    "identified by n=":          6,
    "partners":                  8,
    "neutral M mass":            9,
    "compound_name":            28,
    "adduct":                    9,
    "score":                     7,
    "precursor_mz":              9,
    "ion_mobility":              7,
    "ccs":                       6,
    "rt":                        6,
    "mol_formula":              12,
    "smiles":                   10,
    "inchi":                     8,
    "inchi_key":                16,
    "method":                    8,
}
PEAK_AREA_WIDTH    = 10  # peak area columns
REL_INTENSITY_WIDTH = 8  # relative intensity columns

# --- Load data ---
quant = pd.read_csv(QUANT_PATH)
annot = pd.read_csv(ANNOT_PATH).drop(columns=[c for c in DROP_COLS if c in pd.read_csv(ANNOT_PATH).columns])

sample_cols = [c for c in quant.columns if "mzML" in c]
meta_cols   = [c for c in quant.columns if c not in sample_cols and c != "Unnamed: 20"]

# --- Merge quant + annotations (inner join = annotated features only) ---
annot_sorted = annot.sort_values("score", ascending=False)
merged = quant.merge(annot_sorted, left_on="row ID", right_on="id", how="inner")

# --- RT-based isomer grouping ---
# For each compound_name:
#   1. Group row IDs within RT_TOL of each other (adduction/dimerization artifacts)
#      -> keep the feature with the best score, sum peak areas across redundant features
#   2. If multiple distinct RT groups exist -> true isomers
#      -> rename as compound_name_iso1, compound_name_iso2... in ascending RT order

records = []

for compound, grp in merged.groupby("compound_name"):
    # One row per row ID — keep best score if multiple adducts on the same row ID
    best_per_rowid = grp.sort_values("score", ascending=False).drop_duplicates(subset="row ID")
    best_per_rowid = best_per_rowid.sort_values("row retention time").reset_index(drop=True)

    # RT clustering: split when gap between consecutive features exceeds RT_TOL
    best_per_rowid["rt_group"] = 0
    g = 0
    for i in range(1, len(best_per_rowid)):
        if best_per_rowid.loc[i, "row retention time"] - best_per_rowid.loc[i-1, "row retention time"] > RT_TOL:
            g += 1
        best_per_rowid.loc[i, "rt_group"] = g

    n_groups = best_per_rowid["rt_group"].nunique()

    for rt_group_id, rt_grp in best_per_rowid.groupby("rt_group"):
        # Best feature in the group (highest score)
        best_row = rt_grp.sort_values("score", ascending=False).iloc[0].copy()

        # Sum peak areas across all redundant features in the group
        for col in sample_cols:
            best_row[col] = rt_grp[col].sum()

        # Rename if multiple isomer groups
        if n_groups > 1:
            best_row["compound_name"] = f"{compound}_iso{rt_group_id + 1}"

        records.append(best_row)

final = pd.DataFrame(records).reset_index(drop=True)

# --- Relative intensity columns (max across samples = 1 per compound) ---
rel_cols = []
for col in sample_cols:
    rel_col = col.replace("Peak area", "Rel. intensity")
    row_max = final[sample_cols].max(axis=1)
    final[rel_col] = (final[col] / row_max.replace(0, pd.NA)).round(4)
    rel_cols.append(rel_col)

# --- Column order: meta | annotations | peak areas | relative intensities ---
annot_cols = [c for c in annot_sorted.columns if c != "id" and c not in DROP_COLS]
col_order  = meta_cols + annot_cols + sample_cols + rel_cols
col_order  = [c for c in col_order if c in final.columns]
final = final[col_order]

print(f"Unique annotated features (after grouping): {len(final)}")
print(f"  of which renamed isomers: {final['compound_name'].str.contains('_iso').sum()}")

# --- Export Excel ---
final.to_excel(OUTPUT_PATH, index=False, engine="openpyxl")

# --- Formatting helpers ---
def score_to_hex(score):
    """Pastel: salmon at 0 -> light yellow at 0.5 -> light green at 1."""
    """Red (FF0000) at 0 -> Yellow (FFFF00) at 0.5 -> Green (00CC00) at 1."""
    score = max(0.0, min(1.0, float(score)))
    if score <= 0.5:
        t = score / 0.5
        r, g, b = 255, int(255 * t), 0
    else:
        t = (score - 0.5) / 0.5
        r, g, b = int(255 * (1 - t)), 204, 0
    return f"{r:02X}{g:02X}{b:02X}"

def intensity_to_hex(value):
    """Pastel: light blue (200,220,255) at 0 -> white (255,255,255) at 0.5 -> light red (255,200,200) at 1."""
    value = max(0.0, min(1.0, float(value)))
    if value >= 0.5:
        # White -> pastel red (255,180,180)
        t = (value - 0.5) / 0.5
        r = 255
        g = int(255 - (255 - 180) * t)
        b = int(255 - (255 - 180) * t)
    else:
        # Pastel blue (180,200,255) -> white
        t = value / 0.5
        r = int(180 + (255 - 180) * t)
        g = int(200 + (255 - 200) * t)
        b = 255
    return f"{r:02X}{g:02X}{b:02X}"

# --- Formatting ---
wb = load_workbook(OUTPUT_PATH)
ws = wb.active

header_fill = PatternFill("solid", fgColor="2E4057")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
thin = Side(style="thin", color="CCCCCC")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Map column indices
col_group_fill  = {}
score_col_idx   = None
rel_col_indices = set()
peak_col_indices = set()

for col_idx, col_name in enumerate(final.columns, start=1):
    if col_name == "score":
        score_col_idx = col_idx
    if col_name in rel_cols:
        rel_col_indices.add(col_idx)
    if col_name in sample_cols:
        peak_col_indices.add(col_idx)
    for group, hex_color in GROUP_COLORS.items():
        if group in col_name:
            col_group_fill[col_idx] = PatternFill("solid", fgColor=hex_color)
            break

# Format header row
for col_idx, cell in enumerate(ws[1], start=1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

ws.row_dimensions[1].height = 50

# Format data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        col_idx = cell.column
        cell.font = Font(name="Arial", size=8)
        cell.alignment = Alignment(vertical="center")
        cell.border = cell_border

        if col_idx == score_col_idx and cell.value is not None:
            try:
                cell.fill = PatternFill("solid", fgColor=score_to_hex(cell.value))
            except (ValueError, TypeError):
                pass
        elif col_idx in rel_col_indices and cell.value is not None:
            try:
                cell.fill = PatternFill("solid", fgColor=intensity_to_hex(cell.value))
            except (ValueError, TypeError):
                pass
        elif col_idx in col_group_fill:
            cell.fill = col_group_fill[col_idx]

# Set column widths
for col_idx, col_name in enumerate(final.columns, start=1):
    col_letter = get_column_letter(col_idx)
    if col_name in sample_cols:
        ws.column_dimensions[col_letter].width = PEAK_AREA_WIDTH
    elif col_name in rel_cols:
        ws.column_dimensions[col_letter].width = REL_INTENSITY_WIDTH
    else:
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 10)

ws.freeze_panes = "A2"
wb.save(OUTPUT_PATH)
print(f"Output: {OUTPUT_PATH}")